import openpyxl
import datetime
import os
import itertools

cutoff = 0.5

# rules = ["Date Time", "Pregnancies_flag" ,"Pregnancies_lb","Pregnancies_ub","Glucose_flag","Glucose_lb","Glucose_ub","BP_flag","BP_lb","BP_ub","SkinThickness_flag","SkinThickness_lb","SkinThickness_ub","Insulin_flag","Insulin_lb","Insulin_ub","BMI_flag","BMI_lb","BMI_ub","DPF_flag","DPF_lb","DPF_ub","Age_flag","Age_lb","Age_ub","Class","Hit Ratio"]

def log(attr,accuracy,file_name):
    file_name = "Logs/" + file_name + ".xlsx"

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    now = datetime.datetime.now()
    current_date_time = now.strftime("%Y-%m-%d %H:%M:%S")
    
    for i in range(len(attr)):
        if i % 3 == 0:
            if attr[i] >= cutoff:
                attr[i] = 1
            else:
                attr[i] = 0
    for i in range(1,len(attr),3):
        mx = max(attr[i],attr[i + 1])
        mn = min(attr[i],attr[i + 1])
        attr[i] = mn
        attr[i + 1] = mx
        
    sheet.append([current_date_time] + attr + [accuracy])
    workbook.save(file_name)

def log_confusion_mat(mat):
    file_path = "Logs/confusion.xlsx"
    mat = list(itertools.chain(*mat))
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.append(mat)
    workbook.save(file_path)

def log_MIR(len_0,len_1):
    file_path = "Logs/MIR.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.append([len_0,len_1])
    workbook.save(file_path)
        